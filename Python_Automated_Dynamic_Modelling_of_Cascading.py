import sys  # import sys module in order to call sys.path.append()
sys.path.append("C:\Program Files\DIgSILENT\PowerFactory 2026\Python\3.12")  # adding path to the PowerFactory - module
import powerfactory as pf  # importing the pf-module
app = pf.GetApplication()   # calling app Application object
import pandas as pd
import My_func as rf  # importing the Resilience_functions module
import Excel as xl
import csv
import numpy as np
import openpyxl
import pandas as pd
import math
import statistics


O = app.GetActiveProject()  # getting the active project only to check if there is a project active
if not O: 
    app.PrintError('No project active !!!')
    sys.exit()
app.ResetCalculation()   # start always from the same PowerFactory state
app.ClearOutputWindow()   # clear Output Window (OPTIONAL)

#get all objects/elements within the network
Lines = app.GetCalcRelevantObjects("*.ElmLne")
Buses = app.GetCalcRelevantObjects("*.ElmTerm")
Loads = app.GetCalcRelevantObjects("*.ElmLod")
Cubs = app.GetCalcRelevantObjects("*.StaCubic")
Switches = app.GetCalcRelevantObjects("*.StaSwitch")
Relays = app.GetCalcRelevantObjects("*.ElmRelay")
Transformers = app.GetCalcRelevantObjects("*.ElmTr2")
SynchGens = app.GetCalcRelevantObjects("*.ElmSym")
StatGens = app.GetCalcRelevantObjects("*.ElmGenstat")


Ldf = app.GetFromStudyCase('ComLdf')   # Get commands of calculating load flow
Init = app.GetFromStudyCase('ComInc')  # Get commands of calculating initial conditions
Sim = app.GetFromStudyCase('ComSim')   # Get commands of running simulations
ElmRes_TVFS = app.GetFromStudyCase('TVFS.ElmRes')  # Create class of result variables named "Results"
ComRes = app.GetFromStudyCase('ComRes')  # Get commands of export results
Events_folder = app.GetFromStudyCase('IntEvt')  # Get events folder

def CorrectNames(data):
# Correct the line names in the first column of the data array
    for row_index in range(data.shape[0]):
        object_name = data[row_index, 0]
        # Replace the non-breaking space with a regular space
        corrected_name = object_name.replace('\xa0', ' ')
        data[row_index, 0] = corrected_name
    return data

def CreateResultsFile(objects_list, ElmRes):
    # Delete only IntMon objects with the same name
    for obj_del in ElmRes.GetContents('*.IntMon'):
        obj_del.Delete()
    
    for objects in objects_list:    
        for obj in objects:
            # Create new IntMon for this object
            obj_res = ElmRes.CreateObject('IntMon', f'{obj.loc_name}')
            obj_res.obj_id = obj

            # Assign variables based on object type
            if obj.GetClassName() == 'ElmLne':
                obj_res.vars = ['m:i:bus1:A','m:phii:bus1:A']
            elif obj.GetClassName() == 'ElmTr2':
                obj_res.vars = ['m:i:bushv', 'm:phii:bushv']
            elif obj.GetClassName() == 'ElmTerm':
                obj_res.vars = ['m:u', 'm:phiu', 'm:fe']
            elif obj.GetClassName() in ('ElmSym', 'ElmGenstat'):
                obj_res.vars = ['n:u:bus1:A', 'n:phiu:bus1:A', 's:xspeed',
                                'm:i:bus1:A','m:phii:bus1:A',
                                'm:P:bus1:A', 'm:Q:bus1:A']
            elif obj.GetClassName() == 'ElmLod':
                obj_res.vars = ['m:P:bus1:A', 'm:Q:bus1:A']

def extract_TVFS_results(ElmRes_TVFS):
    Init = app.GetFromStudyCase('ComInc')  # Get commands of calculating initial conditions
    Sim = app.GetFromStudyCase('ComSim')   # Get commands of running simulations
    # Load results
    ElmRes_TVFS.Load()
    NrRow = ElmRes_TVFS.GetNumberOfRows()
    NrCol = ElmRes_TVFS.GetNumberOfColumns()

    # Generate the time vector (NrRow points between start and stop)
    time_values = [ElmRes_TVFS.GetValue(row_idx, -1)[1] for row_idx in range(NrRow)]

    # All monitored objects
    objects = [obj.loc_name for obj in ElmRes_TVFS.GetContents('*.IntMon')]

    # Dictionary of DataFrames
    dataframes = {}

    for obj in objects:
        variables = []
        col_indices = []

        # Find all variables belonging to this object
        for col_idx in range(NrCol):
            obj_type = ElmRes_TVFS.GetObject(col_idx)
            if obj_type is not None and obj_type.loc_name == obj:
                var_name = ElmRes_TVFS.GetVariable(col_idx)
                variables.append(var_name)
                col_indices.append(col_idx)

        # Build data dict (time + variables)
        data = {"time [s]": time_values}
        for var_name, col_idx in zip(variables, col_indices):
            col_values = [ElmRes_TVFS.GetValue(row_idx, col_idx)[1] for row_idx in range(NrRow)]
            data[var_name] = col_values

        # Store DataFrame
        df = pd.DataFrame(data)
        dataframes[obj_type.GetClassName() + obj] = df

    #     # Write all bus_dfs to one Excel file, each in its own sheet
    # with pd.ExcelWriter("TVFS.xlsx", engine="openpyxl") as writer:
    #     for bus_name, df in dataframes.items():
    #         df.to_excel(writer, sheet_name=bus_name[:31], index=False)  

    return dataframes

    
def AddOvercurrentRelay(Network, objects, FrameTyp, LogicDsl, data):
    for obj in objects:
        if obj.GetClassName() == 'ElmLne':
            cub1 = obj.bus1
        elif obj.GetClassName() == 'ElmTr2':
            cub1 = obj.bushv
        ExistRelay = Network.GetContents('*.ElmComp')
        for relay in ExistRelay:
            if relay.loc_name == f'{obj.loc_name} Protection Frame':
                relay.Delete()
        RelayFrame = Network.CreateObject('ElmComp', f'{obj.loc_name} Protection Frame')
        RelayFrame.typ_id = FrameTyp
        Imea = RelayFrame.CreateObject('StaImea', f'Imea {obj.loc_name}')
        Imea.pcubic = cub1; Imea.i_mode = 1 
        # Logic = RelayFrame.CreateObject('ElmDsl', f'Protection {line.loc_name}')
        Logic = RelayFrame.AddCopy(LogicDsl, f'Protection {obj.loc_name}')
        for i in range(data.shape[0]):
                if obj.loc_name == data[i, 0]:
                    Logic.params = data[i, 3:10].tolist()
        # Event = RelayFrame.CreateObject('ElmDsl', f'Tripping {line.loc_name}')
        # Event = RelayFrame.AddCopy(EventDsl, f'Tripping {line.loc_name}')
        RelayFrame.pelm = [Imea, Logic, obj]

def AddSheddingRelay(Network, objects, FrameTyp, LogicDsl, data):
    for obj in objects:
        bus1 = obj.bus1
        ExistRelay = Network.GetContents('*.ElmComp')
        for relay in ExistRelay:
            if relay.loc_name == f'{obj.loc_name} Protection Frame':
                relay.Delete()
        RelayFrame = Network.CreateObject('ElmComp', f'{obj.loc_name} Protection Frame')
        RelayFrame.typ_id = FrameTyp
        Vmea = RelayFrame.CreateObject('StaVmea', f'Vmea {obj.loc_name}')
        Vmea.pbusbar = bus1; Vmea.i_mode = 1 
        # Logic = RelayFrame.CreateObject('ElmDsl', f'Protection {line.loc_name}')
        Logic = RelayFrame.AddCopy(LogicDsl, f'Protection {obj.loc_name}')
        for i in range(data.shape[0]):
            if obj.loc_name == data[i, 0]:
                if obj.GetClassName() == 'ElmLod':
                    Logic.params = data[i, 3:26].tolist()
                elif obj.GetClassName() == 'ElmSym' or obj.GetClassName() == 'ElmGenstat':
                    Logic.params = data[i, 3:22].tolist()
        # Event = RelayFrame.CreateObject('ElmDsl', f'Tripping {line.loc_name}')
        # Event = RelayFrame.AddCopy(EventDsl, f'Tripping {load.loc_name}')
        RelayFrame.pelm = [Vmea, Logic, obj]

def ChangeLoadType(LdTyps) :       
    for item in LdTyps:
        # app.PrintPlain(item)
        item.aP = 0.391
        item.bP = 0.42
        item.bQ = -1

def LoadStep(Loads,step:float):
    for load in Loads:
        load.scale=1
    for load in Loads:
        load.scale+=step

def GetCombinations(objects, data):
    import itertools
    comb = []; combs = []
    prob1 = 0; prob2 = 0; prob = []
    for obj in itertools.combinations(objects, 2) :
        comb.append(obj[0])
        comb.append(obj[1])
        combs.append(comb)
        for i in range(data.shape[0]):
            if obj[0].loc_name == data[i, 0] :
                prob1 = float(data[i, -1])
            if obj[1].loc_name == data[i, 0] :
                prob2 = float(data[i, -1])
        if prob1*prob2 == 0:
            app.PrintPlain('Probability calculation ERROR')
        prob.append(prob1*prob2)
        comb= []
        prob1 = 0
        prob2 = 0
    return combs, prob

def weighted_unique_choices(data, weights, num):
    if len(data) != len(weights):
        raise ValueError("Data and weights must be the same length.")
    if num > len(data):
        raise ValueError("Cannot select more unique items than available in data.")
    weights = np.array(weights, dtype=float)
    if np.any(weights < 0) or np.sum(weights) == 0:
        raise ValueError("Weights must be non-negative and sum to a positive number.")
    probabilities = weights / np.sum(weights)
    indices = np.random.choice(len(data), size=num, replace=False, p=probabilities)
    return [data[i] for i in indices]

def weighted_choice(data, weights):
    import random
    if not data:
        raise ValueError("Data list cannot be empty.")

    # Calculate the cumulative sums of the weights
    total_weight = sum(weights)
    if total_weight <= 0:
        raise ValueError("Weights must sum to a positive number.")

    cumulative_weights = []
    current_sum = 0
    for w in weights:
        current_sum += w
        cumulative_weights.append(current_sum)

    # Generate a random number between 0 and the total weight
    rand_num = random.uniform(0, total_weight)

    # Find which bucket the random number falls into
    for i, cum_weight in enumerate(cumulative_weights):
        if rand_num < cum_weight:
            return data[i], weights[i]

    # Fallback to the last element in case of floating point inaccuracies
    return data[-1], weights[-1]

def create_cases_list(cases, prob):
    list_cases = []
    for i in range(len(cases)):
        a, b = None, None
        for item in Lines:
            if item == cases[i][0]:
                a = item.loc_name
            if item == cases[i][1]:
                b = item.loc_name
        list_cases.append([i, a, b, prob[i]])
    return list_cases

def compute_pairwise_rms(TVFS_results, outage_time, tstop, fnom, features=["m:phiu", "m:fe"]):
    bus_names = []
    diffs_dict = {feat: [] for feat in features}

    # --- Step 1: Prepare differences per bus and feature ---
    for name, df in TVFS_results.items():
        if "ElmTerm" not in name:
            continue

        bus_names.append(name)
        # Subset after outage
        subset = df.loc[df["time [s]"] >= outage_time, ["time [s]"] + features].copy()

        for feat in features:
            col_name = f"{name}_{feat}"

            # Reference value at t=0
            ref_value = df.loc[df["time [s]"] == 0, feat].iloc[0]
            if feat == "m:fe":
                ref_value *= 2 * np.pi * fnom  # convert reference to rad/s
            elif feat == "m:phiu":
                ref_value = np.deg2rad(ref_value)  # convert reference to radians

            # Subset values
            data = subset[feat].copy()
            if feat == "m:fe":
                data = data * 2 * np.pi * fnom  # convert subset to rad/s
            elif feat == "m:phiu":
                data = np.deg2rad(data)  # convert subset to radians

            # Differences from reference
            df_diff = pd.DataFrame({
                "time [s]": subset["time [s]"],
                col_name: data - ref_value
            })
            diffs_dict[feat].append(df_diff)

    # for feat, dfs in diffs_dict.items():
    #     for k, df in enumerate(dfs):
    #         app.PrintPlain(f"{feat} {k} shape={df.shape} dtypes={df.dtypes.to_dict()}")
    #         app.PrintPlain(df.head().to_string())

    # --- Step 2: Merge all buses per feature ---
    merged_diffs = {}
    for feat in features:
        dfs = diffs_dict[feat]
        if not dfs:
            continue

        # Standardize time column
        for i in range(len(dfs)):
            dfs[i] = dfs[i].copy()
            dfs[i]["time [s]"] = dfs[i]["time [s]"].round(6)
            dfs[i].set_index("time [s]", inplace=True)

        # Take the time grid from the first bus
        ref_time = dfs[0].index

        # Reindex all dfs to that same grid (with interpolation)
        aligned_dfs = []
        for df in dfs:
            df = df.reindex(ref_time).interpolate(method="linear")
            aligned_dfs.append(df)

        # Concatenate side by side
        df_merged = pd.concat(aligned_dfs, axis=1)

        merged_diffs[feat] = df_merged

    # --- Step 3: Compute pairwise differences and sum squared ---
    time_array = merged_diffs[features[0]].index.values
    mask = (time_array >= outage_time) & (time_array <= tstop)
    time_window = time_array[mask]

    n_buses = len(bus_names)
    pairs = [(i, j) for i in range(n_buses) for j in range(i+1, n_buses)]
    pair_names = [(bus_names[i], bus_names[j]) for i, j in pairs]

    squared_sum = np.zeros((len(pairs), len(time_window)))

    for feat in features:
        arr = merged_diffs[feat].values
        pairwise_diffs = np.array([arr[:, i] - arr[:, j] for i, j in pairs])
        squared_sum += pairwise_diffs[:, mask]**2

    # --- Step 4: Trapezoidal integration and RMS ---
    integrals = np.trapz(squared_sum, x=time_window, axis=1)
    rms_values = np.sqrt(integrals / (tstop - outage_time))

    # --- Step 5: Return results ---
    rms_results = dict(zip(pair_names, rms_values))
    return rms_results

def dissimilarity_matrix(rms_results):
    import numpy as np

    # Extract bus names from rms_results keys
    bus_names = sorted({bus for pair in rms_results.keys() for bus in pair})
    n_buses = len(bus_names)

    # Initialize distance matrix
    dissimilarity_matrix = np.zeros((n_buses, n_buses))

    # Fill distance matrix
    for (bus_i, bus_j), rms in rms_results.items():
        i = bus_names.index(bus_i)
        j = bus_names.index(bus_j)
        dissimilarity_matrix[i, j] = rms
        dissimilarity_matrix[j, i] = rms  # symmetric
    # Remove class names of buses
    bus_names = [name.replace("ElmTerm", "") for name in bus_names]

    return dissimilarity_matrix, bus_names

def cluster_buses(dissimilarity_matrix, bus_names, n_clusters=None):
    from scipy.cluster.hierarchy import linkage, fcluster
    from scipy.spatial.distance import squareform
    import math

    n_buses = len(bus_names)
    assert dissimilarity_matrix.shape == (n_buses, n_buses), \
    "Dissimilarity matrix must be square and match number of bus names"

    # Convert to condensed form for linkage
    condensed_dist = squareform(dissimilarity_matrix)

    # Hierarchical clustering using Ward method
    Z = linkage(condensed_dist, method='ward')

    # Thumb rule: number of clusters = sqrt(number of buses)
    if n_clusters is None:
        n_clusters = max(2, round(math.sqrt(n_buses)))  # at least 2 clusters

    # Cut dendrogram at threshold
    clusters = fcluster(Z, t=n_clusters, criterion='maxclust')

    # Map buses to clusters using original fcluster IDs
    temp_groups = {}
    for bus, cluster_id in zip(bus_names, clusters):
        temp_groups.setdefault(cluster_id, []).append(bus)

    # Remap to simple consecutive integers in ascending order of cluster_id
    coherent_groups = {}
    for new_id, old_id in enumerate(sorted(temp_groups.keys()), start=1):
        coherent_groups[new_id] = temp_groups[old_id]

    return coherent_groups

def TSI(coherent_groups, SynchGens, TVFS_results, delta_lim):
    # --- Precompute generator inertia weights ---
    Project = app.GetActiveProject()
    sys_base = Project.pPrjSettings.Sbase

    gen_info = {}
    for gen in SynchGens:
        gen_typ = gen.typ_id
        M_i = 2 * gen_typ.h * (gen_typ.sgn / sys_base)
        bus_name = gen.bus1.cterm.loc_name
        key = f"ElmTerm{bus_name}"
        if key not in TVFS_results:
            continue
        gen_info[gen.loc_name] = {
            "M_i": M_i,
            "df_angles": TVFS_results[key][["m:phiu"]]
        }

    if not gen_info:
        return ["Ungenerated System"] * len(coherent_groups)

    # --- Compute total COI (vectorized sum) ---
    Mtotal = sum(info["M_i"] for info in gen_info.values())
    weighted_dfs = [info["df_angles"] * info["M_i"] for info in gen_info.values()]
    df_sum = pd.concat(weighted_dfs, axis=1).sum(axis=1)
    df_coi = (df_sum / Mtotal).to_frame(name="delta_COI")

    # --- Vectorized computation of COI-referred bus angles ---
    # Collect all "ElmTerm" bus angle DataFrames
    bus_dfs = {}
    for key, df in TVFS_results.items():
        if key.startswith("ElmTerm") and "m:phiu" in df:
            bus_name = key.replace("ElmTerm", "")
            bus_dfs[bus_name] = df["m:phiu"]

    # Concatenate into one DataFrame: index = time, columns = buses
    if bus_dfs:
        df_all = pd.concat(bus_dfs, axis=1)  # Multi-column DataFrame with bus names

        # Subtract COI in one vectorized operation
        df_all_coi_ref = df_all.sub(df_coi["delta_COI"], axis=0)

    # df_all_coi_ref.to_csv("angles_coi_ref.csv", index=True)

    
    # --- Precompute generator M_i mapping for fast lookup ---
    gen_M_map = {gen.loc_name: 2 * gen.typ_id.h * (gen.typ_id.sgn / sys_base) for gen in SynchGens}
    bus_to_gen = {gen.bus1.cterm.loc_name: gen.loc_name for gen in SynchGens}

    # --- Compute TSI per cluster ---
    TSI_values = []
    cluster_coi_series = {}  # Store time series for each cluster

    for cluster_id, buses in coherent_groups.items():
        # Identify valid generators in this cluster
        valid_gens = [bus_to_gen[bus] for bus in buses if bus in bus_to_gen]
        if not valid_gens:
            TSI_values.append("Ungenerated Group")
            cluster_coi_series[cluster_id] = None
            continue

        # Compute inertia weight mapping for this cluster
        Mcluster = sum(gen_M_map[gen] for gen in valid_gens)

        # Get bus names corresponding to these generators
        cluster_buses = [bus for bus in buses if bus in bus_to_gen]

        # Weighted sum of COI-referenced angles
        weighted_series = []
        for bus in cluster_buses:
            gen = bus_to_gen[bus]
            M_i = gen_M_map[gen]
            weighted_series.append(df_all_coi_ref[bus] * M_i)

        # Sum across buses
        df_sum_cluster = pd.concat(weighted_series, axis=1).sum(axis=1)

        # Normalize by cluster inertia
        df_cluster_coi = (df_sum_cluster / Mcluster).to_frame(name=f"delta_cluster_COI_{cluster_id}")

        # Store time series for CSV export
        cluster_coi_series[cluster_id] = df_cluster_coi

        # --- Compute TSI ---
        max_angle = df_cluster_coi[f"delta_cluster_COI_{cluster_id}"].abs().max()
        if pd.isna(max_angle):
            TSI_cluster = "Ungenerated Group"
        elif max_angle <= delta_lim:
            TSI_cluster = 0.0
        elif max_angle >= 180.0:
            TSI_cluster = 1.0
        else:
            TSI_cluster = (max_angle - delta_lim) / (180.0 - delta_lim)

        TSI_values.append(TSI_cluster)
    app.PrintPlain(TSI_values)

    # # --- Export all cluster COI time series to CSV ---
    # # Combine all clusters into one DataFrame
    # combined_df = pd.concat(cluster_coi_series.values(), axis=1)
    # combined_df.to_csv("cluster_coi_timeseries.csv", index=True)
        
    return TSI_values
        

def max_sequential_violation(df, i_min, i_max, t_max):
    time = df["time [s]"].to_numpy()
    values = df["m:u"].to_numpy()
    
    # Condition mask
    mask = (values < i_min) | (values > i_max) & (values != 0)
    
    max_duration = 0.0
    start_time = None
    
    for i, flag in enumerate(mask):
        if flag:
            if start_time is None:  # entering a violation
                start_time = time[i]
        else:
            if start_time is not None:  # exiting a violation
                duration = time[i-1] - start_time
                max_duration = max(max_duration, duration)
                start_time = None
    
    # Handle case where it ends inside a violation
    if start_time is not None:
        duration = time[-1] - start_time
        max_duration = max(max_duration, duration)
    
    return min(max_duration/t_max,1)

def cluster_max(values_dict, buses):
    cluster_vals = [values_dict[b] for b in buses if b in values_dict]
    return max(cluster_vals) if cluster_vals else None  # or 0.0 if you prefer

def VDI(coherent_groups, TVFS_results, u_min=0.9, u_max=1.1, tu_max=60):
    # Collect all "ElmTerm" bus angle DataFrames
    bus_dfs = {}
    VDI = {}
    for key, df in TVFS_results.items():
        if key.startswith("ElmTerm") and "m:u" in df:
            bus_name = key.replace("ElmTerm", "")
            bus_dfs[bus_name] = df[["time [s]", "m:u"]]
            VDI[bus_name] = max_sequential_violation(bus_dfs[bus_name], u_min, u_max, tu_max)
    # --- Compute VDI per cluster ---
    VDI_values = [cluster_max(VDI, buses) for buses in coherent_groups.values()]

    app.PrintPlain(VDI_values)
    return VDI_values

def avg_rocof_over_window(time_vec, rocof_vec, window_s):
    mask_window = time_vec <= time_vec[0] + window_s
    if np.any(mask_window):
        return np.mean(np.abs(rocof_vec[mask_window]))
    else:
        return np.nan  # no data in window

def max_nadir_rocof(TVFS_results, DB, delta_f_max, rocof_max, t_short, Fnom=50):
    time = TVFS_results["time [s]"].to_numpy()
    values = TVFS_results["m:fe"].to_numpy()*Fnom # Convert to Hz
    f_dev = np.abs(values[values != 0] - Fnom)  # Exclude zero values and get absolute deviation from nominal
    f_nadir = np.max(f_dev[f_dev>=DB]) if f_dev[f_dev>=DB].size > 0 else 0.0  # Nadir-based FDI with a deadband

    time_post = time[(values != 0) & (time >= t_short)]
    values_post = values[(values != 0) & (time >= t_short)]
    rocof = np.gradient(values_post, time_post)
    
    avg_rocof_100ms = avg_rocof_over_window(time_post, rocof, 0.1)
    avg_rocof_500ms = avg_rocof_over_window(time_post, rocof, 0.5)
        
    return min(f_nadir/delta_f_max,1), min(avg_rocof_100ms/rocof_max,1), min(avg_rocof_500ms/rocof_max,1)


def FDI(coherent_groups, TVFS_results, t_short, Fnom=50, DB=0.1, delta_f_max=2.5, rocof_max=2.5):
    # Collect all "ElmTerm" bus angle DataFrames
    bus_dfs = {}
    FDI = {}; RCFI_100 = {}; RCFI_500 = {}
    for key, df in TVFS_results.items():
        if key.startswith("ElmTerm") and "m:u" in df:
            bus_name = key.replace("ElmTerm", "")
            bus_dfs[bus_name] = df[["time [s]", "m:fe"]]
            FDI[bus_name], RCFI_100[bus_name], RCFI_500[bus_name]= max_nadir_rocof(bus_dfs[bus_name], DB, delta_f_max, rocof_max, t_short, Fnom)

    FDI_values       = [cluster_max(FDI, buses) for buses in coherent_groups.values()]
    RCFI_100_values  = [cluster_max(RCFI_100, buses) for buses in coherent_groups.values()]
    RCFI_500_values  = [cluster_max(RCFI_500, buses) for buses in coherent_groups.values()]

    app.PrintPlain(FDI_values), app.PrintPlain(RCFI_100_values), app.PrintPlain(RCFI_500_values)
    return FDI_values, RCFI_100_values, RCFI_500_values

def RunN2Outages(num, list_cases, Events_folder, delta_lim, sim_time, Network):
    import time
    import itertools
    time_start = time.time()
    start = 0
    end = num
    results = []
    CreateResultsFile([Buses],ElmRes_TVFS)
    Init.p_resvar = ElmRes_TVFS

    for i in range(start, end) :

        EventSet = Events_folder.GetContents()
        for event in EventSet:
            event.outserv=1  # switch off all outage events
        app.PrintPlain(i)
        app.PrintPlain(list_cases[i][1:3])
        # Init.Execute()

        for item in Lines:
            item.outserv = 0  # put all lines in service

        for item in Lines:
            ## If you want to include line outages uncomment the following lines
            # if item.loc_name == list_cases[i][1]:
            #     Outage1 = Events_folder.CreateObject('EvtSwitch', f'{item.loc_name} Outage Event 2')
            #     Outage1.p_target = item
            #     Outage1.time = 0.05  # starts at t= 0.05s
            #     Outage1.i_switch = 0  # take the element out of service
            ## If you want to start with an N-1 case uncomment the following lines
            if item.loc_name == list_cases[i][1]:
                item.outserv = 1 # take N-1 element out of service
            if item.loc_name == list_cases[i][2]:
                Short = Events_folder.CreateObject('EvtShc', f'{item.loc_name} Short ckt Event 2')
                Short.p_target = item # event target is the each line
                Short.time = 0.00  # starts at t= 0s
                Short.i_shc = 0  # 3-phase short circuit
                Short.R_f = 0; Short.X_f = 0  # zero impedance
            if item.loc_name == list_cases[i][2]:
                Outage2 = Events_folder.CreateObject('EvtSwitch', f'{item.loc_name} Short clearing Event 2')
                Outage2.p_target = item # event target is the each line
                Outage2.time = 0.05  # starts at t= 0.05s
                Outage2.i_switch = 0  # take the element out of service
        
        Ldf.iopt_net=0
        Init.Execute()
        Sim.tstop = sim_time # simulation time

        Sim.Execute()
        
        EventSet = Events_folder.GetContents()
        UVLS=0; OVLS=0; OFLS=0; UFLS=0; RoCoF_pos=0; RoCoF_neg=0; TotalShed=0
        for event in EventSet:
            if event.loc_name[:4] == 'UVLS':
                load = event.p_target
                UVLS += load.plini
            elif event.loc_name[:4] == 'OVLS':
                load = event.p_target
                OVLS += load.plini
            elif event.loc_name[:4] == 'UFLS':
                load = event.p_target
                UFLS += load.plini 
            elif event.loc_name[:4] == 'OFLS':
                load = event.p_target
                OFLS += load.plini
            elif event.loc_name[:8] == 'RoCoF+LS':
                load = event.p_target
                RoCoF_pos += load.plini
            elif event.loc_name[:8] == 'RoCoF-LS':
                load = event.p_target
                RoCoF_neg += load.plini
        TotalShed = UVLS + OVLS + UFLS + OFLS + RoCoF_pos + RoCoF_neg
        time_end = time.time()
        app.PrintPlain(f'Total Load Shed: {TotalShed} MW')

        TVFS_results = extract_TVFS_results(ElmRes_TVFS)
        features = ["m:phiu", "m:fe"]  # Can add more features if needed
        rms_results = compute_pairwise_rms(TVFS_results, Outage2.time, Sim.tstop, Network.frnom, features=features)
        # Step 2: Dissimilarity matrix
        D, bus_names = dissimilarity_matrix(rms_results)
        # Step 3: Cluster buses
        coherent_groups = cluster_buses(D, bus_names)

        app.PrintPlain(coherent_groups)
        T_index = TSI(coherent_groups, SynchGens, TVFS_results, delta_lim)
        V_index = VDI(coherent_groups, TVFS_results, u_min=0.9, u_max=1.1, tu_max=60)
        F_index, RCF100_index, RCF500_index = FDI(coherent_groups, TVFS_results, Short.time, Network.frnom, DB=0.1, delta_f_max=2.5, rocof_max=2.5)
        results.append([i, list_cases[i][1], list_cases[i][2], time_end - time_start,TotalShed, UVLS, OVLS, UFLS, OFLS, RoCoF_pos, RoCoF_neg, str(coherent_groups),*T_index, *V_index, *F_index, *RCF100_index, *RCF500_index])
    return results, coherent_groups

# -----------------------------
# Helper: approximate t critical for 95% CI
# -----------------------------
def t_critical_95(df):
    """
    Returns two-sided 95% t critical value for small df.
    For df > 30, 1.96 is a good approximation.
    """
    t_table = {
        1: 12.706, 2: 4.303, 3: 3.182, 4: 2.776, 5: 2.571,
        6: 2.447, 7: 2.365, 8: 2.306, 9: 2.262, 10: 2.228,
        11: 2.201, 12: 2.179, 13: 2.160, 14: 2.145, 15: 2.131,
        16: 2.120, 17: 2.110, 18: 2.101, 19: 2.093, 20: 2.086,
        21: 2.080, 22: 2.074, 23: 2.069, 24: 2.064, 25: 2.060,
        26: 2.056, 27: 2.052, 28: 2.048, 29: 2.045, 30: 2.042
    }
    if df <= 0:
        return None
    if df <= 30:
        return t_table[df]
    return 1.96


# -----------------------------
# Helper: compute current CI stats
# -----------------------------
def compute_ci_halfwidth(samples_y):
    """
    samples_y: list of Y_k = probability_k * load_shed_k
    returns: (mean_y, halfwidth)
    """
    n = len(samples_y)

    if n < 2:
        return None, None

    mean_y = statistics.mean(samples_y)
    s = statistics.stdev(samples_y)
    tcrit = t_critical_95(n - 1)
    halfwidth = tcrit * s / math.sqrt(n)

    return mean_y, halfwidth


LoadTypes = app.GetCalcRelevantObjects("*.TypLod")
ChangeLoadType(LoadTypes)
LoadStep(Loads, 0) # Initial load step in pu

# Network = app.GetCalcRelevantObjects("Nine-bus System.ElmNet")[0] # Change accordingly to the grid name in your project Case9
Network = app.GetCalcRelevantObjects("Grid.ElmNet")[0] # Change accordingly to the grid name in your project Case39

UserFolder = app.GetProjectFolder("blk")


# # Add Line OC relays
# LineRelayFolder = UserFolder.GetChildren(0,"Line Relay*")[0]
# LineRelayFrameTyp = LineRelayFolder.GetChildren(0,'Line Relay.BlkDef')[0]
# LogicTyp = LineRelayFolder.GetChildren(0,'Line protection.BlkDef')[0]; LogicDsl = LogicTyp.GetChildren(0,'*.ElmDsl')[0]
# # data_Lines = xl.read_excel('case9.xlsx','Lines', 1, 7, 1, 17, 1) # case9
# data_Lines = xl.read_excel('case39.xlsx','Lines', 1, 35, 1, 17, 1) # case39
# data_Lines = CorrectNames(data_Lines)
# AddOvercurrentRelay(Network, Lines, LineRelayFrameTyp, LogicDsl, data_Lines)

# # Add Transformer OC relays
# TrafoRelayFolder = UserFolder.GetChildren(0,"Transformer Relay*")[0]
# TrafoRelayFrameTyp = TrafoRelayFolder.GetChildren(0,'TX Relay.BlkDef')[0]
# LogicTyp = TrafoRelayFolder.GetChildren(0,'TX protection.BlkDef')[0]; LogicDsl = LogicTyp.GetChildren(0,'*.ElmDsl')[0]
# # data_Trafos = xl.read_excel('case9.xlsx','Trafos', 1, 4, 1, 16, 1) # case9
# data_Trafos = xl.read_excel('case39.xlsx','Trafos', 1, 13, 1, 16, 1) # case39
# data_Trafos = CorrectNames(data_Trafos)
# AddOvercurrentRelay(Network, Transformers, TrafoRelayFrameTyp, LogicDsl, data_Trafos)

# # Add Load O-UV, O-UF, +-RoCoF relays
# LoadRelayFolder = UserFolder.GetChildren(0,"Load Relay*")[0]
# LoadRelayFrameTyp = LoadRelayFolder.GetChildren(0,'LS Relay.BlkDef')[0]
# LogicTyp = LoadRelayFolder.GetChildren(0,'LS protection.BlkDef')[0]; LogicDsl = LogicTyp.GetChildren(0,'*.ElmDsl')[0]
# # data_Loads = xl.read_excel('case9.xlsx','Loads', 1, 10, 1, 26, 1) # case9
# data_Loads = xl.read_excel('case39.xlsx','Loads', 1, 20, 1, 26, 1) # case39
# data_Loads = CorrectNames(data_Loads)
# AddSheddingRelay(Network, Loads, LoadRelayFrameTyp, LogicDsl, data_Loads)

# # Add Generators O-V/HZ, O-UF, +-RoCoF relays    
# GenRelayFolder = UserFolder.GetChildren(0,"Generator Relay*")[0]
# GenRelayFrameTyp = GenRelayFolder.GetChildren(0,'GS Relay.BlkDef')[0]
# LogicTyp = GenRelayFolder.GetChildren(0,'GS protection.BlkDef')[0]; LogicDsl = LogicTyp.GetChildren(0,'*.ElmDsl')[0]
# # data_SynchGens = xl.read_excel('case9.xlsx','SynchGens', 1, 3, 1, 22, 1) # case9
# data_SynchGens = xl.read_excel('case39.xlsx','SynchGens', 1, 11, 1, 22, 1) # case39
# data_SynchGens = CorrectNames(data_SynchGens)
# AddSheddingRelay(Network, SynchGens, GenRelayFrameTyp, LogicDsl, data_SynchGens)

# data_out = xl.read_excel('case9.xlsx','Lines', 1, 7, 1, 17, 1) # case9
data_out = xl.read_excel('case39.xlsx','Lines', 1, 35, 1, 17, 1) # case39
data_out = CorrectNames(data_out)
combs, prob = GetCombinations(Lines, data_out)

delta_lim = 60 # Maximum angle in deg
sim_time = 60 # Simulation time in s

# # randomly choose
# num = 100 # number of cases to choose

# # cases = weighted_unique_choices(combs, prob, num) # To run only unique cases
# samples = [weighted_choice(combs, prob) for _ in range(num)]
# cases, probs_cases = zip(*samples)
# list_cases = create_cases_list(cases, probs_cases)


# results, coherent_groups = RunN2Outages(num, list_cases, Events_folder, delta_lim, sim_time, Network)
# Window = app.GetOutputWindow()

# # Create a new workbook and select the active worksheet
# workbook = openpyxl.Workbook()
# sheet = workbook.active
# sheet.title = f'Random Cases {Network.loc_name}'


# cluster_ids = list(coherent_groups.keys())
# # Write headers to the first row (optional but good practice)
# sheet.cell(row=1, column=1, value='Index')
# sheet.cell(row=1, column=2, value='Line 1')
# sheet.cell(row=1, column=3, value='Line 2')
# sheet.cell(row=1, column=4, value='Probability')
# sheet.cell(row=1, column=5, value='Total Simulation Time (s)')
# sheet.cell(row=1, column=6, value='Total Load Shed (MW)')
# sheet.cell(row=1, column=7, value='UVLS (MW)')
# sheet.cell(row=1, column=8, value='OVLS (MW)')
# sheet.cell(row=1, column=9, value='UFLS (MW)')
# sheet.cell(row=1, column=10, value='OFLS (MW)')
# sheet.cell(row=1, column=11, value='RoCoF+LS (MW)')
# sheet.cell(row=1, column=12, value='RoCoF-LS  (MW)')
# sheet.cell(row=1, column=13, value='Coherent Groups Dictionary')

# num_clusters = len(cluster_ids)
# start_col = 14  # starting column
# row = 1
# for i, cluster_id in enumerate(cluster_ids):
#     sheet.cell(row=row, column=start_col + i, value=f'TSI: {cluster_id}')
#     sheet.cell(row=row, column=start_col + i + num_clusters, value=f'VDI: {cluster_id}')
#     sheet.cell(row=row, column=start_col + i + num_clusters*2, value=f'FDI: {cluster_id}')
#     sheet.cell(row=row, column=start_col + i + num_clusters*3, value=f'RCFI_100: {cluster_id}')
#     sheet.cell(row=row, column=start_col + i + num_clusters*4, value=f'RCFI_500: {cluster_id}')

# # Write data to the spreadsheet
# # Note: openpyxl uses 1-based indexing for rows and columns
# for i in range(len(cases)):
#     results[i].insert(3, probs_cases[i])
#     # The row number needs to be i + 2 to account for the 1-based index and header row
#     for j in range(len(results[i])):
#         sheet.cell(row=i + 2, column=j+1, value=results[i][j])
# # Save the workbook
# workbook.save('Random Cases.xlsx')

# -----------------------------
# Sequential MC parameters
# -----------------------------
batch_size = 5          # run a few cases at a time
min_cases = 30          # do not stop too early
max_cases = 5000        # safety cap
rel_tol = 0.01          # 1%
abs_tol = 1e-6          # fallback if mean is ~0

all_cases = []
all_probs = []
all_results = []
all_y = []
coherent_groups = {}

total_done = 0

while total_done < max_cases:
    # Sample a batch
    samples = [weighted_choice(combs, prob) for _ in range(batch_size)]
    cases_batch, probs_batch = zip(*samples)
    list_cases_batch = create_cases_list(cases_batch, probs_batch)

    # Run this batch
    results_batch, coherent_groups = RunN2Outages(
        len(cases_batch),
        list_cases_batch,
        Events_folder,
        delta_lim,
        sim_time,
        Network
    )

    # Store results and build Y_k = p_k * LS_k
    for i in range(len(cases_batch)):
        p_case = probs_batch[i]
        raw_result = results_batch[i]

        # IMPORTANT:
        # Assuming raw_result BEFORE inserting probability is ordered as:
        # [Index, Line 1, Line 2, Total Simulation Time, Total Load Shed, UVLS, ...]
        # so Total Load Shed is raw_result[4]
        load_shed = raw_result[4]

        yk = p_case * load_shed

        all_cases.append(cases_batch[i])
        all_probs.append(p_case)
        all_results.append(raw_result)
        all_y.append(yk)

    total_done += len(cases_batch)

    # Check stopping criterion after enough samples
    if total_done >= min_cases:
        mean_y, halfwidth = compute_ci_halfwidth(all_y)

        if mean_y is not None:
            threshold = max(rel_tol * abs(mean_y), abs_tol)

            print(f"n = {total_done}, mean = {mean_y:.6f}, halfwidth = {halfwidth:.6f}, threshold = {threshold:.6f}")

            if halfwidth < threshold:
                print("Stopping criterion reached.")
                break

Window = app.GetOutputWindow()
Window.PrintPlain(f"Sequential MC finished with {total_done} cases.")
if len(all_y) >= 2:
    mean_y, halfwidth = compute_ci_halfwidth(all_y)
    Window.PrintPlain(f"Estimated expected load shed = {mean_y:.6f}")
    Window.PrintPlain(f"95% CI half-width = {halfwidth:.6f}")
    Window.PrintPlain(f"Relative half-width = {100.0 * halfwidth / max(abs(mean_y), 1e-12):.4f}%")
else:
    Window.PrintPlain("Not enough samples to compute confidence interval.")


# -----------------------------
# Write results to Excel
# -----------------------------
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = f'Random Cases {Network.loc_name}'

cluster_ids = list(coherent_groups.keys())

sheet.cell(row=1, column=1, value='Index')
sheet.cell(row=1, column=2, value='Line 1')
sheet.cell(row=1, column=3, value='Line 2')
sheet.cell(row=1, column=4, value='Probability')
sheet.cell(row=1, column=5, value='Total Simulation Time (s)')
sheet.cell(row=1, column=6, value='Total Load Shed (MW)')
sheet.cell(row=1, column=7, value='UVLS (MW)')
sheet.cell(row=1, column=8, value='OVLS (MW)')
sheet.cell(row=1, column=9, value='UFLS (MW)')
sheet.cell(row=1, column=10, value='OFLS (MW)')
sheet.cell(row=1, column=11, value='RoCoF+LS (MW)')
sheet.cell(row=1, column=12, value='RoCoF-LS  (MW)')
sheet.cell(row=1, column=13, value='Coherent Groups Dictionary')
sheet.cell(row=1, column=14, value='p * Load Shed (MW)')

num_clusters = len(cluster_ids)
start_col = 15
row = 1
for i, cluster_id in enumerate(cluster_ids):
    sheet.cell(row=row, column=start_col + i, value=f'TSI: {cluster_id}')
    sheet.cell(row=row, column=start_col + i + num_clusters, value=f'VDI: {cluster_id}')
    sheet.cell(row=row, column=start_col + i + num_clusters * 2, value=f'FDI: {cluster_id}')
    sheet.cell(row=row, column=start_col + i + num_clusters * 3, value=f'RCFI_100: {cluster_id}')
    sheet.cell(row=row, column=start_col + i + num_clusters * 4, value=f'RCFI_500: {cluster_id}')

# Write each stored result
for i in range(len(all_results)):
    row_data = list(all_results[i])  # copy raw result
    row_data.insert(3, all_probs[i]) # insert probability after Line 2
    row_data.insert(13, all_y[i])    # insert p * Load Shed after coherent groups dictionary column

    for j in range(len(row_data)):
        sheet.cell(row=i + 2, column=j + 1, value=row_data[j])

# Summary at bottom
summary_row = len(all_results) + 4
sheet.cell(row=summary_row, column=1, value='Total sampled cases')
sheet.cell(row=summary_row, column=2, value=len(all_results))

if len(all_y) >= 2:
    mean_y, halfwidth = compute_ci_halfwidth(all_y)
    sheet.cell(row=summary_row + 1, column=1, value='Estimated expected load shed')
    sheet.cell(row=summary_row + 1, column=2, value=mean_y)

    sheet.cell(row=summary_row + 2, column=1, value='95% CI half-width')
    sheet.cell(row=summary_row + 2, column=2, value=halfwidth)

    sheet.cell(row=summary_row + 3, column=1, value='Relative half-width (%)')
    sheet.cell(row=summary_row + 3, column=2, value=100.0 * halfwidth / max(abs(mean_y), 1e-12))

workbook.save('Random Cases.xlsx')

